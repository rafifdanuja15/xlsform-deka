import os
import io
import uuid
import logging
import time
import subprocess
from pathlib import Path
from flask import Flask, request, jsonify, send_file, render_template
from flask_cors import CORS
from werkzeug.utils import secure_filename
from dotenv import load_dotenv

load_dotenv()

logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s")
logger = logging.getLogger(__name__)

app = Flask(
    __name__,
    template_folder="../frontend/templates",
    static_folder="../frontend/static",
)
CORS(app)

UPLOAD_FOLDER = Path(__file__).parent.parent / "tmp_uploads"
UPLOAD_FOLDER.mkdir(exist_ok=True)

ALLOWED_EXTENSIONS = {"pdf", "doc", "docx"}
MAX_FILE_SIZE = 10 * 1024 * 1024  # 10MB


def allowed_file(filename: str) -> bool:
    return "." in filename and filename.rsplit(".", 1)[1].lower() in ALLOWED_EXTENSIONS


# ── Helper: konversi .doc → .docx via LibreOffice/antiword ──────────────────
def _convert_doc_to_docx(save_path: Path) -> tuple[Path, bool, str]:
    """
    Coba konversi file .doc ke .docx.
    Return: (path_hasil, berhasil, metode)
      - path_hasil: Path ke .docx jika berhasil, tetap save_path jika gagal
      - berhasil: True jika konversi sukses
      - metode: 'libreoffice' | 'antiword_text' | 'failed'
    """
    docx_path = save_path.with_suffix(".docx")

    # Kandidat path LibreOffice
    custom_path = os.environ.get("SOFFICE_PATH", "")
    soffice_candidates = list(filter(None, [
        custom_path,
        "soffice",
        "libreoffice",
        r"C:\Program Files\LibreOffice\program\soffice.exe",
        r"C:\Program Files (x86)\LibreOffice\program\soffice.exe",
        r"C:\Program Files\LibreOffice 7\program\soffice.exe",
        r"C:\Program Files\LibreOffice 24\program\soffice.exe",
        r"C:\Program Files\LibreOffice 25\program\soffice.exe",
        r"D:\Program Files\LibreOffice\program\soffice.exe",
        r"D:\Program Files (x86)\LibreOffice\program\soffice.exe",
        r"D:\Program Files\LibreOffice 7\program\soffice.exe",
        r"D:\Program Files\LibreOffice 24\program\soffice.exe",
        r"D:\Program Files\LibreOffice 25\program\soffice.exe",
    ]))

    for soffice in soffice_candidates:
        try:
            result = subprocess.run(
                [soffice, "--headless", "--convert-to", "docx",
                 "--outdir", str(save_path.parent), str(save_path)],
                capture_output=True, timeout=90
            )
            if result.returncode == 0 and docx_path.exists():
                logger.info(f".doc dikonversi ke .docx via {soffice}: {docx_path.name}")
                return docx_path, True, "libreoffice"
        except (FileNotFoundError, subprocess.TimeoutExpired, OSError):
            continue

    # Fallback: antiword → teks mentah
    try:
        result = subprocess.run(
            ["antiword", str(save_path)],
            capture_output=True, text=True, timeout=30
        )
        if result.returncode == 0 and result.stdout.strip():
            logger.info(f".doc dibaca via antiword: {len(result.stdout)} chars")
            txt_path = save_path.with_suffix(".txt")
            txt_path.write_text(result.stdout, encoding="utf-8")
            return txt_path, True, "antiword_text"
    except (FileNotFoundError, subprocess.TimeoutExpired, OSError):
        pass

    return save_path, False, "failed"


def _docx_to_html(doc_path: str) -> str:
    """
    Konversi .docx ke HTML agar browser bisa render tabel, bold, italic, dll
    semirip mungkin dengan tampilan Word — tanpa ASCII art.
    """
    import docx as _docx
    from docx.text.paragraph import Paragraph
    from docx.table import Table
    import html

    doc = _docx.Document(doc_path)
    parts = []

    def render_paragraph(para: Paragraph) -> str:
        """Render satu paragraf ke HTML, jaga bold/italic/underline per-run."""
        if not para.text.strip():
            return '<p style="margin:0;line-height:1.4">&nbsp;</p>'

        inner = ''
        for run in para.runs:
            text = html.escape(run.text)
            if not text:
                continue
            if run.bold:
                text = f'<strong>{text}</strong>'
            if run.italic:
                text = f'<em>{text}</em>'
            if run.underline:
                text = f'<u>{text}</u>'
            inner += text

        # Style heading berdasarkan style name Word
        style_name = (para.style.name or '').lower()
        if 'heading 1' in style_name:
            return f'<h2 style="font-size:15px;font-weight:700;margin:8px 0 4px">{inner}</h2>'
        elif 'heading' in style_name:
            return f'<h3 style="font-size:13px;font-weight:700;margin:6px 0 3px">{inner}</h3>'
        else:
            return f'<p style="margin:0 0 2px;line-height:1.5">{inner or "&nbsp;"}</p>'

    def render_cell(cell) -> str:
        """Render isi cell — bisa multi-paragraf."""
        cell_html = ''
        for para in cell.paragraphs:
            cell_html += render_paragraph(para)
        return cell_html

    def render_table(tbl: Table) -> str:
        rows_html = ''
        for ri, row in enumerate(tbl.rows):
            # Deduplikasi sel merged (python-docx duplikasi sel merged)
            seen_ids = set()
            cells_html = ''
            for cell in row.cells:
                cid = id(cell._tc)
                if cid in seen_ids:
                    continue
                seen_ids.add(cid)
                bg = '#f0f4ff' if ri == 0 else 'transparent'
                fw = '600' if ri == 0 else 'normal'
                cells_html += (
                    f'<td style="border:1px solid #d1d5db;padding:5px 8px;'
                    f'vertical-align:top;background:{bg};font-weight:{fw};'
                    f'font-size:11px;min-width:80px">'
                    f'{render_cell(cell)}</td>'
                )
            rows_html += f'<tr>{cells_html}</tr>'
        return (
            '<div style="overflow-x:auto;margin:8px 0">'
            '<table style="border-collapse:collapse;width:100%;font-size:11px">'
            f'{rows_html}</table></div>'
        )

    body = doc.element.body
    for child in body:
        tag = child.tag.split('}')[-1] if '}' in child.tag else child.tag
        if tag == 'p':
            para = Paragraph(child, doc)
            parts.append(render_paragraph(para))
        elif tag == 'tbl':
            tbl = Table(child, doc)
            parts.append(render_table(tbl))

    return '\n'.join(parts)


@app.route("/")
def index():
    return render_template("index.html")


@app.route("/health")
def health():
    return jsonify({"status": "ok", "version": "2.0.0"})

@app.route("/api/preview-xlsform/<uid>")
def preview_xlsform(uid: str):
    """
    Baca file xlsx hasil konversi dan kembalikan isinya sebagai JSON untuk preview.
    Dibatasi 50 baris pertama per sheet; total_rows berisi jumlah baris asli.
    """
    PREVIEW_LIMIT = 50
 
    matches = list(UPLOAD_FOLDER.glob(f"{uid}_*"))
    if not matches:
        return jsonify({"error": "File tidak ditemukan atau sudah kadaluarsa"}), 404
 
    out_path = matches[0]
    try:
        import openpyxl
        wb = openpyxl.load_workbook(str(out_path), read_only=True)
 
        def _sheet_to_rows(ws):
            headers = []
            rows = []
            total_rows = 0
            for ri, row in enumerate(ws.iter_rows(values_only=True)):
                vals = [str(v) if v is not None else "" for v in row]
                if ri == 0:
                    headers = vals
                else:
                    if any(v.strip() for v in vals):
                        total_rows += 1
                        if len(rows) < PREVIEW_LIMIT:
                            rows.append(dict(zip(headers, vals)))
            return {"headers": headers, "rows": rows, "total_rows": total_rows}
 
        survey_data  = _sheet_to_rows(wb["survey"])  if "survey"  in wb.sheetnames else {"headers": [], "rows": [], "total_rows": 0}
        choices_data = _sheet_to_rows(wb["choices"]) if "choices" in wb.sheetnames else {"headers": [], "rows": [], "total_rows": 0}
        wb.close()
 
        return jsonify({"survey": survey_data, "choices": choices_data})
    except Exception as e:
        logger.error(f"Preview xlsform gagal: {e}", exc_info=True)
        return jsonify({"error": f"Gagal membaca hasil konversi: {str(e)[:200]}"}), 500

@app.route("/api/preview-xlsform/<uid>/search")
def preview_xlsform_search(uid: str):
    """
    Cari baris spesifik di hasil konversi berdasarkan name (survey) atau list_name (choices).
    Query params:
      - name      : filter sheet survey by kolom 'name'
      - list_name : filter sheet choices by kolom 'list_name'
    Return semua baris yang match (tanpa limit).
    """
    filter_name      = request.args.get("name", "").strip()
    filter_list_name = request.args.get("list_name", "").strip()

    if not filter_name and not filter_list_name:
        return jsonify({"error": "Parameter 'name' atau 'list_name' wajib diisi"}), 400

    matches = list(UPLOAD_FOLDER.glob(f"{uid}_*"))
    if not matches:
        return jsonify({"error": "File tidak ditemukan atau sudah kadaluarsa"}), 404

    out_path = matches[0]
    try:
        import openpyxl
        wb = openpyxl.load_workbook(str(out_path), read_only=True)

        def _filter_sheet(ws, filter_col, filter_val):
            headers = []
            rows = []
            col_idx = None
            for ri, row in enumerate(ws.iter_rows(values_only=True)):
                vals = [str(v) if v is not None else "" for v in row]
                if ri == 0:
                    headers = vals
                    # Cari index kolom yang jadi filter
                    for ci, h in enumerate(headers):
                        if h.strip().lower() == filter_col.lower():
                            col_idx = ci
                            break
                else:
                    if col_idx is not None and vals[col_idx].strip() == filter_val:
                        rows.append(dict(zip(headers, vals)))
            return {"headers": headers, "rows": rows, "total_rows": len(rows)}

        result = {}
        if filter_name and "survey" in wb.sheetnames:
            result["survey"] = _filter_sheet(wb["survey"], "name", filter_name)
        if filter_list_name and "choices" in wb.sheetnames:
            result["choices"] = _filter_sheet(wb["choices"], "list_name", filter_list_name)

        wb.close()
        return jsonify(result)

    except Exception as e:
        logger.error(f"Search preview gagal: {e}", exc_info=True)
        return jsonify({"error": f"Gagal mencari data: {str(e)[:200]}"}), 500

@app.route("/api/download/<uid>")
def download_file(uid: str):
    """Endpoint untuk mengunduh file hasil konversi berdasarkan UID."""
    matches = list(UPLOAD_FOLDER.glob(f"{uid}_*"))
    if not matches:
        return jsonify({"error": "File tidak ditemukan atau sudah kadaluarsa"}), 404
    out_path    = matches[0]
    output_name = out_path.name[len(uid) + 1:]
    return send_file(
        str(out_path),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=output_name,
    )


@app.route("/api/convert", methods=["POST"])
def convert():
    if "file" not in request.files:
        return jsonify({"error": "Tidak ada file di request"}), 400

    file = request.files["file"]
    if not file.filename:
        return jsonify({"error": "Tidak ada file yang dipilih"}), 400

    if not allowed_file(file.filename):
        return jsonify({"error": "Format tidak didukung. Gunakan PDF, DOC, atau DOCX."}), 400

    file.seek(0, os.SEEK_END)
    file_size = file.tell()
    file.seek(0)
    if file_size > MAX_FILE_SIZE:
        return jsonify({"error": "File terlalu besar. Maksimal 10MB."}), 400

    filename  = secure_filename(file.filename)
    uid       = str(uuid.uuid4())[:8]
    save_path = UPLOAD_FOLDER / f"{uid}_{filename}"

    try:
        file.save(str(save_path))
        logger.info(f"File disimpan: {save_path} ({file_size:,} bytes)")
        t0  = time.time()
        ext = filename.rsplit(".", 1)[1].lower()

        # ── Konversi .doc legacy ke .docx ────────────────────────────────────
        if ext == "doc":
            converted_path, ok, method = _convert_doc_to_docx(save_path)

            if method == "antiword_text" and ok:
                logger.warning(".doc → antiword text — fallback ke LLM pipeline")
                xlsx_bytes = _llm_pipeline(str(converted_path))
                converted_path.unlink(missing_ok=True)
                elapsed     = time.time() - t0
                logger.info(f"Selesai dalam {elapsed:.1f}s | output {len(xlsx_bytes):,} bytes")
                output_name = filename.rsplit(".", 1)[0] + "_xlsform.xlsx"
                out_uid     = str(uuid.uuid4())[:8]
                out_path    = UPLOAD_FOLDER / f"{out_uid}_{output_name}"
                out_path.write_bytes(xlsx_bytes)
                return jsonify({
                    "download_uid":  out_uid,
                    "download_name": output_name,
                    "notes": {"fallback_questions": [], "placeholder_choices": []},
                })

            if not ok:
                return jsonify({
                    "error": (
                        "File .doc (format Word lama) tidak bisa dikonversi otomatis. "
                        "Silakan buka file di Microsoft Word atau LibreOffice, lalu simpan ulang sebagai .docx, "
                        "kemudian upload file .docx tersebut."
                    )
                }), 400

            save_path = converted_path
            ext       = "docx"

        if ext in ("doc", "docx"):
            logger.info("Mode: Deka Research DOCX parser")
            try:
                from .docx_parser import parse_questionnaire
                from .json_to_xlsform import convert_json_to_xlsform

                questions = parse_questionnaire(str(save_path))
                parsed    = {"_meta": {"source": filename}, "questions": questions}
                logger.info(f"Parse selesai: {len(questions)} baris")

                xlsx_bytes, conversion_notes = convert_json_to_xlsform(parsed)

            except ValueError as e:
                logger.warning(f"Bukan format Deka: {e} — fallback ke LLM pipeline")
                xlsx_bytes        = _llm_pipeline(str(save_path))
                conversion_notes  = {"fallback_questions": [], "placeholder_choices": []}

        elif ext == "pdf":
            logger.info("Mode: LLM pipeline (PDF)")
            xlsx_bytes       = _llm_pipeline(str(save_path))
            conversion_notes = {"fallback_questions": [], "placeholder_choices": []}

        else:
            return jsonify({"error": "Format tidak didukung"}), 400

        elapsed = time.time() - t0
        logger.info(f"Selesai dalam {elapsed:.1f}s | output {len(xlsx_bytes):,} bytes")

        output_name = filename.rsplit(".", 1)[0] + "_xlsform.xlsx"
        out_uid     = str(uuid.uuid4())[:8]
        out_path    = UPLOAD_FOLDER / f"{out_uid}_{output_name}"
        out_path.write_bytes(xlsx_bytes)

        return jsonify({
            "download_uid":  out_uid,
            "download_name": output_name,
            "notes":         conversion_notes,
        })

    except ValueError as e:
        logger.error(f"Validation error: {e}")
        return jsonify({"error": str(e)}), 422
    except Exception as e:
        logger.error(f"Konversi gagal: {e}", exc_info=True)
        return jsonify({"error": f"Konversi gagal: {str(e)[:300]}"}), 500
    finally:
        if save_path.exists():
            save_path.unlink(missing_ok=True)


def _llm_pipeline(file_path: str) -> bytes:
    """Fallback pipeline lama via LLM untuk PDF atau format non-Deka."""
    from .file_parser import parse_uploaded_file
    from .llm_client import call_llm_for_xlsform
    from .xlsform_builder import build_xlsform_from_json

    text = parse_uploaded_file(file_path)
    if not text or len(text.strip()) < 50:
        raise ValueError("Tidak dapat membaca konten file.")

    xlsform_data = call_llm_for_xlsform(text)
    return build_xlsform_from_json(xlsform_data)


if __name__ == "__main__":
    port  = int(os.environ.get("PORT", 5000))
    debug = os.environ.get("FLASK_ENV", "production") == "development"
    app.run(host="0.0.0.0", port=port, debug=debug)
