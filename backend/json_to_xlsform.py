"""
json_to_xlsform.py — Deterministik + LLM hanya untuk skip_logic & tipe ambigu
=============================================================================
Fix dari audit:
- Pisahkan instruksi DP dari label → hint
- required=no untuk pertanyaan probe/OE optional
- select_multiple: gunakan selected() untuk relevant, bukan = 
- constraint hanya untuk STOP di screening (SA), bukan semua
- LLM: kirim sub-pertanyaan lengkap (a1a/a1c, v4a, dll) agar relevant valid
- Validasi: semua list_name yang direferensikan harus ada di choices
"""

from __future__ import annotations

import io, json, logging, os, re, httpx
from collections import Counter
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openai import OpenAI

logger = logging.getLogger(__name__)

SUMOPOD_API_KEY  = os.environ.get("SUMOPOD_API_KEY", "")
SUMOPOD_BASE_URL = os.environ.get("SUMOPOD_BASE_URL", "https://ai.sumopod.com/v1")
SUMOPOD_MODEL    = os.environ.get("SUMOPOD_MODEL", "claude-sonnet-4-6")
API_TIMEOUT      = int(os.environ.get("API_TIMEOUT", "120"))

_ROUTE_TO_TYPE: dict[str, str] = {
    "SA": "select_one", "MA": "select_multiple",
    "OE": "text", "M (OE)": "text", "MA+OE": "select_multiple",
    "SA+MA": "select_one", "SA+MA+MA": "select_one",
    "info": "note", "begin_group": "begin_group", "end_group": "end_group",
    "begin_repeat": "begin_repeat", "end_repeat": "end_repeat",
    "integer": "integer",
    # Tipe literal dari blok info_responden
    "text": "text", "date": "date", "time": "time",
    "decimal": "decimal", "geopoint": "geopoint",
}

_TYPE_KEYWORDS: list[tuple[re.Pattern, str]] = [
    (re.compile(r'\btanggal lahir\b|\bdate of birth\b', re.I), "date"),
    (re.compile(r'\busia aktual\b|\bumur anda\b', re.I),       "integer"),
    (re.compile(r'\bjam\b|\bwaktu\b|\btime\b', re.I),          "time"),
]

# Pertanyaan OE yang tidak required
_OPTIONAL_PATTERNS = re.compile(
    r'\bapalagi\b|\bprobe\b|\bselain\b|\blainnya\b|\bsebutkan\b|\balasan\b'
    r'|\bmengapa\b|\bpendapat\b|\bsaran\b', re.I
)

_STOP_MESSAGE = "Mohon maaf, Anda tidak memenuhi kriteria responden untuk survei ini."

# Instruksi yang harus dipindah ke hint (bukan label)
_LABEL_CLEANUP_RE = re.compile(
    r'\b(TUNJUKKAN KARTU BANTU|KARTU BANTU|SPONTAN|PROBE|ROTASIKAN|'
    r'BACAKAN|INTERVIEWER|DP\s*:|NOTE\s*:|TANYAKAN\s+\w+\s+JIKA\b)'
    r'[^\n]*',
    re.I
)


def _safe_name(raw_id: str) -> str:
    name = raw_id.strip().lower()
    name = re.sub(r'[^a-z0-9_]', '_', name)
    name = re.sub(r'_+', '_', name).strip('_')
    if name and name[0].isdigit():
        name = 'q_' + name
    return name or 'unknown'


def _clean_label(text: str) -> str:
    """Hapus instruksi DP dari label — sisakan hanya pertanyaan untuk responden."""
    # Hapus baris yang dimulai dengan marker instruksi
    lines = text.splitlines()
    clean = []
    for ln in lines:
        ln = ln.strip()
        if not ln:
            continue
        up = ln.upper()
        if any(up.startswith(m) for m in [
            "TUNJUKKAN", "KARTU BANTU", "DP:", "NOTE:", "[DP",
            "TANYAKAN ", "DITANYAKAN", "SPONTAN", "PROBE", "BACAKAN",
            "ROTASIKAN", "TRANSFER JAWABAN", "MASUKAN LIST",
        ]):
            continue
        if re.match(r'^\[.*\]$', ln):
            continue
        clean.append(ln)
    return " ".join(clean).strip() or text.strip()


def _make_client() -> OpenAI:
    return OpenAI(
        api_key=SUMOPOD_API_KEY,
        base_url=SUMOPOD_BASE_URL,
        timeout=httpx.Timeout(API_TIMEOUT),
        max_retries=0,
    )


# ── LLM: skip_logic + tipe ambigu ─────────────────────────────────────────────

def _llm_resolve(ambiguous: list[dict], skip_logic_qs: list[dict],
                 all_q_ids: list[str]) -> dict[str, dict]:
    """Satu batch LLM request untuk resolve skip_logic + tipe ambigu."""
    if not SUMOPOD_API_KEY:
        logger.warning("[LLM] API key tidak ada — gunakan fallback deterministik")
        return {}
    if not ambiguous and not skip_logic_qs:
        return {}

    client = _make_client()
    items = []

    for q in ambiguous:
        items.append({
            "task": "determine_type",
            "id": q["id"],
            "question": q["question"][:150],
            "route_type": q["route_type"],
            "has_choices": bool(q.get("choices")),
            "num_choices": len(q.get("choices", [])),
        })

    for q in skip_logic_qs:
        items.append({
            "task": "translate_skip_logic",
            "id": q["id"],
            "skip_logic": q["skip_logic"][:250],
            "question": q["question"][:100],
        })

    # Berikan konteks ID yang tersedia di survey
    id_context = ", ".join(all_q_ids[:80])

    prompt = f"""Kamu adalah ahli XLSForm KoboToolbox ODK. Proses tugas berikut.

ID pertanyaan yang tersedia di survey (gunakan untuk relevant):
{id_context}

TUGAS:
{json.dumps(items, ensure_ascii=False, indent=2)}

INSTRUKSI untuk "determine_type":
- Type valid: select_one, select_multiple, text, integer, decimal, date, note, calculate
- has_choices=true + route SA → select_one; route MA → select_multiple
- has_choices=false + route OE → text
- Usia/angka/harga → integer; tanggal → date
- Pertanyaan SEC/header tanpa choices → note

INSTRUKSI untuk "translate_skip_logic":
- Terjemahkan ke ODK XPath expression untuk kolom "relevant"
- Nama variabel: konversi ID ke snake_case (A1c → a1c, Q1 → q1, V4a → v4a)
- PENTING: gunakan selected() untuk select_multiple, = untuk select_one
  Contoh select_one:   ${{s5}} = '1'
  Contoh select_multiple: selected(${{q1}}, '7')
- "TERKODE X di Y" → cek tipe Y: jika MA gunakan selected(${{y}}, 'X'), jika SA gunakan ${{y}} = 'X'
- Referensi ke ID yang ADA di daftar di atas saja
- Jika tidak bisa diterjemahkan dengan yakin → kembalikan ""

OUTPUT: JSON array, satu object per item:
[
  {{"id": "Q7b", "relevant": "selected(${{a1c}}, '7') and not(selected(${{q3}}, '7'))"}},
  {{"id": "V14", "relevant": "${{v4a}} = '5'"}},
  {{"id": "S9",  "type": "select_one"}}
]

Output HANYA JSON array, tanpa markdown."""

    logger.info(f"[LLM] Mengirim {len(items)} item ({len(ambiguous)} ambigu, {len(skip_logic_qs)} skip_logic)")

    try:
        resp = client.chat.completions.create(
            model=SUMOPOD_MODEL, max_tokens=4096, temperature=0.0,
            messages=[{"role": "user", "content": prompt}],
        )
        raw = resp.choices[0].message.content.strip()
        raw = re.sub(r'^```json\s*', '', raw, flags=re.MULTILINE)
        raw = re.sub(r'^```\s*', '', raw, flags=re.MULTILINE)
        raw = re.sub(r'\s*```$', '', raw, flags=re.MULTILINE)
        idx = raw.find('[')
        if idx > 0: raw = raw[idx:]
        results = json.loads(raw)
        logger.info(f"[LLM] Resolved {len(results)} items")
        return {r["id"]: r for r in results if "id" in r}
    except Exception as e:
        err = str(e)
        if "budget_exceeded" in err or "Budget has been exceeded" in err:
            logger.warning("[LLM] ⚠️  Budget Sumopod habis — top-up di https://sumopod.com/dashboard/ai/keys")
        elif "429" in err:
            logger.warning("[LLM] ⚠️  Rate limit Sumopod")
        else:
            logger.error(f"[LLM] Gagal: {e}")
        return {}


# ── Core converter ─────────────────────────────────────────────────────────────

def convert_json_to_xlsform(parsed_json: dict) -> tuple[bytes, dict]:
    questions: list[dict] = parsed_json.get("questions", [])
    meta: dict            = parsed_json.get("_meta", {})
    logger.info(f"Memproses {len(questions)} baris dari {meta.get('source','?')}")

    # Metadata konversi untuk dilaporkan ke frontend
    conversion_notes: dict = {
        "fallback_questions": [],   # pertanyaan yg pakai fallback type
        "placeholder_choices": [],  # pertanyaan yg list-nya hilang → placeholder
    }

    # Kumpulkan semua ID yang valid (untuk konteks LLM)
    all_q_ids = [
        _safe_name(q["id"]) for q in questions
        if q.get("id") and not q.get("is_header")
        and q.get("route_type","") not in ("begin_group","end_group","end_repeat","begin_repeat")
    ]

    # Identifikasi ambigu & skip_logic
    ambiguous, skip_logic_qs = [], []
    for q in questions:
        if q.get("is_header") or not q.get("id"):
            continue
        rt = q.get("route_type", "").strip()
        has_choices = bool(q.get("choices"))
        if rt == "" and q["id"] and not _detect_type_from_text(q.get("question",""), has_choices):
            ambiguous.append(q)
        elif rt not in _ROUTE_TO_TYPE and rt not in ("", "info"):
            ambiguous.append(q)
        if q.get("skip_logic"):
            skip_logic_qs.append(q)

    logger.info(f"Ambigu: {len(ambiguous)} | skip_logic: {len(skip_logic_qs)}")

    llm_results = _llm_resolve(ambiguous, skip_logic_qs, all_q_ids)
    logger.info(f"[LLM] Resolved: {len(llm_results)}")

    # Build
    survey_rows:  list[dict] = []
    choices_rows: list[dict] = []
    used_names:   dict[str, int] = {}
    current_section = None

    # ── Pre-inject list_kota dari metadata yang disisipkan parser ─────────────
    for q in questions:
        if q.get("route_type") == "kota_choices_data" and q.get("choices"):
            for ch in q["choices"]:
                code  = str(ch.get("code", "")).strip()
                label = str(ch.get("label", "")).strip()
                if code and label:
                    cname = re.sub(r'[^a-z0-9]', '_', label.lower())
                    cname = re.sub(r'_+', '_', cname).strip('_') or f"kota_{code}"
                    choices_rows.append({"list_name": "list_kota", "name": cname, "label": label})
            break  # hanya satu entry kota_choices_data

    # Metadata
    for mt in ("start", "end", "deviceid"):
        survey_rows.append({"type": mt, "name": mt, "label": ""})
        used_names[mt] = 1

    for q in questions:
        q_id       = q.get("id", "")
        is_header  = q.get("is_header", False)
        section    = q.get("section", "")
        route_type = q.get("route_type", "").strip()
        raw_label  = q.get("question", "")
        hint_text  = q.get("hint", "")
        choices    = q.get("choices", [])
        has_stop   = any(c.get("routing") == "STOP" for c in choices)

        # Section group
        if section != current_section and section:
            if current_section is not None:
                grp = f"grp_{_safe_name(current_section[:20])}"
                survey_rows.append({"type": "end_group", "name": grp, "label": ""})
            current_section = section
            grp = f"grp_{_safe_name(section[:20])}"
            used_names[grp] = 1
            survey_rows.append({"type": "begin_group", "name": grp, "label": section})

        # begin_group / end_group dari parser
        if route_type in ("begin_group", "end_group", "begin_repeat", "end_repeat"):
            grp_name = _safe_name(q_id) if q_id else f"grp_{_safe_name(raw_label[:20])}"
            if route_type in ("begin_group", "begin_repeat"):
                if grp_name in used_names:
                    used_names[grp_name] += 1
                    grp_name = f"{grp_name}_{used_names[grp_name]}"
                else:
                    used_names[grp_name] = 1
            survey_rows.append({
                "type": route_type, "name": grp_name,
                "label": _clean_label(raw_label) if route_type.startswith("begin") else "",
            })
            continue

        if is_header:
            # Skip kota_choices_data — sudah diproses sebelum loop
            if route_type == "kota_choices_data":
                continue
            if q_id:
                nn = _safe_name(q_id)
                nn = _dedup_name(nn, used_names)
                survey_rows.append({"type": "note", "name": nn,
                                    "label": _clean_label(raw_label)})
            continue

        if not q_id:
            continue

        name = _safe_name(q_id)
        name = _dedup_name(name, used_names)
        lname = f"list_{name}"

        xlstype = _resolve_type(q, llm_results)

        # Catat pertanyaan yang pakai fallback (tidak ada route_type valid & tidak resolved LLM)
        rt_original = q.get("route_type", "").strip()
        if (
            rt_original not in _ROUTE_TO_TYPE
            and rt_original not in ("", "info", "begin_group", "end_group", "begin_repeat", "end_repeat", "kota_choices_data")
            and q_id not in llm_results
        ) or (
            rt_original == "" and q_id not in llm_results
            and not _detect_type_from_text(q.get("question", ""), bool(choices))
        ):
            conversion_notes["fallback_questions"].append({
                "id": q_id,
                "label": raw_label[:80],
                "route_type": rt_original or "(kosong)",
                "resolved_type": xlstype,
            })

        # Handle tipe lengkap dengan list_name (mis. "select_one list_kota")
        full_type_override = None
        if " " in xlstype and xlstype.split()[0] in ("select_one", "select_multiple"):
            parts = xlstype.split(None, 1)
            full_type_override = xlstype
            xlstype = parts[0]
            lname = parts[1]

        # Bersihkan label
        label = _clean_label(raw_label)
        extra_hint = _LABEL_CLEANUP_RE.findall(raw_label)
        if extra_hint and not hint_text:
            hint_text = " ".join(extra_hint)

        # Relevant dari LLM
        relevant = ""
        if q_id in llm_results and "relevant" in llm_results[q_id]:
            relevant = llm_results[q_id].get("relevant", "") or ""

        # Constraint STOP — hanya untuk screening SA
        constraint = constraint_msg = ""
        if has_stop and xlstype == "select_one" and section.upper() == "SCREENING":
            stop_codes = [c["code"] for c in choices
                          if c.get("routing") == "STOP" and c.get("code")]
            if stop_codes:
                constraint = " and ".join(f". != '{c}'" for c in stop_codes)
                constraint_msg = _STOP_MESSAGE

        # Required: prioritas dari _required (blok info), lalu otomatis
        explicit_required = q.get("_required", "")
        if explicit_required:
            required = explicit_required
        else:
            is_optional = (xlstype == "text" and bool(_OPTIONAL_PATTERNS.search(label)))
            required = "" if xlstype in ("note","calculate","begin_group","end_group") \
                           or is_optional else "yes"

        # Bangun type string final
        if full_type_override:
            type_str = full_type_override
        elif xlstype in ("select_one","select_multiple"):
            type_str = f"{xlstype} {lname}"
        else:
            type_str = xlstype

        row: dict[str, Any] = {"type": type_str, "name": name, "label": label}
        if hint_text:             row["hint"] = hint_text
        if required:              row["required"] = required
        if relevant:              row["relevant"] = relevant
        if constraint:            row["constraint"] = constraint
        if constraint_msg:        row["constraint_message"] = constraint_msg

        survey_rows.append(row)

        # Choices
        if xlstype in ("select_one","select_multiple") and choices:
            seen_names_local: dict[str, int] = {}
            for ch in choices:
                code  = ch.get("code","").strip()
                label_ch = ch.get("label","").strip()
                if not code and not label_ch:
                    continue
                if re.search(r'[A-Za-z]\d+[a-z]?\s*\(', code):
                    continue  # instruksi DP
                if re.match(r'^ROTASIKAN|^\[DP', code, re.I) or \
                   re.match(r'^ROTASIKAN|^\[DP', label_ch, re.I):
                    continue
                raw_cn = re.sub(r'[^a-z0-9]','_', code.lower()) if code \
                         else re.sub(r'[^a-z0-9]','_', label_ch.lower()[:25])
                raw_cn = re.sub(r'_+','_', raw_cn).strip('_') or 'opt'
                # Deduplikasi lokal dalam list ini
                if raw_cn in seen_names_local:
                    seen_names_local[raw_cn] += 1
                    cn = f"{raw_cn}_{seen_names_local[raw_cn]}"
                else:
                    seen_names_local[raw_cn] = 1
                    cn = raw_cn
                choices_rows.append({
                    "list_name": lname, "name": cn, "label": label_ch or code,
                })

    # Tutup group
    if current_section:
        survey_rows.append({
            "type": "end_group",
            "name": f"grp_{_safe_name(current_section[:20])}",
            "label": "",
        })

    # Validasi & placeholder untuk list yang hilang
    defined = {r["list_name"] for r in choices_rows}
    for row in survey_rows:
        t = row.get("type","")
        if t.startswith(("select_one ","select_multiple ")):
            ref = t.split()[1]
            if ref not in defined:
                logger.warning(f"[VALIDATE] list '{ref}' hilang — tambah placeholder")
                choices_rows.append({"list_name": ref, "name": "placeholder",
                                     "label": f"[TODO: isi pilihan {ref}]"})
                defined.add(ref)
                # Cari nama pertanyaan yang mereferensikan list ini
                q_label = next(
                    (r.get("label", "") for r in survey_rows if r.get("type", "") == t),
                    ""
                )
                conversion_notes["placeholder_choices"].append({
                    "list_name": ref,
                    "question_type": t,
                    "label": q_label[:80],
                })

    # Cek duplikat name final
    nc = Counter(r["name"] for r in survey_rows
                 if r.get("name") and not r.get("type","").startswith("end_"))
    dupes = {n:c for n,c in nc.items() if c > 1}
    if dupes:
        logger.error(f"[VALIDATE] Duplicate names: {dupes}")
    else:
        logger.info("[VALIDATE] ✓ Tidak ada duplicate name")

    # Settings
    src  = meta.get("source","kuesioner")
    settings = {
        "form_title":       src.replace(".docx","").replace("_"," ").replace("-"," ").title(),
        "form_id":          re.sub(r'[^a-z0-9_]','_', src.lower().replace(".docx","")),
        "version":          "1",
        "default_language": "Indonesian",
    }

    logger.info(f"Survey rows: {len(survey_rows)} | Choices: {len(choices_rows)}")
    return _build_excel(survey_rows, choices_rows, settings), conversion_notes


def _dedup_name(name: str, used: dict[str, int]) -> str:
    if name in used:
        used[name] += 1
        return f"{name}_{used[name]}"
    used[name] = 1
    return name


def _detect_type_from_text(question_text: str, has_choices: bool) -> str | None:
    for pattern, xlstype in _TYPE_KEYWORDS:
        if pattern.search(question_text):
            return xlstype
    return None


def _resolve_type(q: dict, llm_results: dict) -> str:
    q_id = q.get("id","")
    rt   = q.get("route_type","").strip()
    question = q.get("question","")
    has_choices = bool(q.get("choices"))

    if q_id in llm_results and "type" in llm_results[q_id]:
        return llm_results[q_id]["type"]

    # Handle tipe lengkap dengan list_name (mis. "select_one list_kota")
    if rt.startswith("select_one ") or rt.startswith("select_multiple "):
        return rt  # dikembalikan penuh — akan ditangani di build loop

    if rt in _ROUTE_TO_TYPE:
        mapped = _ROUTE_TO_TYPE[rt]
        if mapped in ("select_one","select_multiple") and not has_choices:
            return mapped  # list akan diisi atau placeholder
        return mapped

    detected = _detect_type_from_text(question, has_choices)
    if detected:
        return detected

    return "select_one" if has_choices else "text"


# ── Excel builder ──────────────────────────────────────────────────────────────

_H_SURVEY   = PatternFill("solid", fgColor="1F4E79")
_H_CHOICES  = PatternFill("solid", fgColor="145A32")
_H_SETTINGS = PatternFill("solid", fgColor="4A235A")
_H_FONT     = Font(color="FFFFFF", bold=True, name="Calibri", size=11)
_GROUP_FILL = PatternFill("solid", fgColor="D6E4F0")
_GROUP_FONT = Font(bold=True, italic=True, name="Calibri", size=10)
_NOTE_FILL  = PatternFill("solid", fgColor="FFF9C4")
_STOP_FILL  = PatternFill("solid", fgColor="FDECEA")
_OPT_FILL   = PatternFill("solid", fgColor="F5F5F5")
_THIN       = Side(style="thin", color="BFBFBF")
_BORDER     = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

_SURVEY_COLS  = ["type","name","label","hint","required","relevant",
                 "constraint","constraint_message","appearance","default","calculation"]
_CHOICES_COLS = ["list_name","name","label"]
_SETTINGS_COLS= ["form_title","form_id","version","default_language",
                 "instance_name","submission_url"]


def _write_header(ws, cols, fill):
    for ci, col in enumerate(cols, 1):
        cell = ws.cell(row=1, column=ci, value=col)
        cell.font = _H_FONT; cell.fill = fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _BORDER
    ws.row_dimensions[1].height = 22


def _build_excel(survey_rows, choices_rows, settings) -> bytes:
    wb   = openpyxl.Workbook()
    ws_s = wb.active;             ws_s.title = "survey"
    ws_c = wb.create_sheet("choices")
    ws_t = wb.create_sheet("settings")

    used_s = [c for c in _SURVEY_COLS
              if c in ("type","name","label") or any(r.get(c) for r in survey_rows)]
    _write_header(ws_s, used_s, _H_SURVEY)

    for ri, row in enumerate(survey_rows, 2):
        rtype = str(row.get("type","")).lower()
        is_group  = rtype.startswith(("begin_group","end_group","begin_repeat","end_repeat"))
        is_note   = rtype == "note"
        is_meta   = rtype in ("start","end","deviceid")
        has_con   = bool(row.get("constraint"))
        is_opt    = row.get("required","") == "" and not is_group and not is_note and not is_meta

        for ci, col in enumerate(used_s, 1):
            val  = row.get(col,"")
            cell = ws_s.cell(row=ri, column=ci, value=str(val) if val else "")
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = _BORDER
            if is_meta:
                cell.font = Font(name="Calibri", size=9, color="888888", italic=True)
            elif is_group:
                cell.fill = _GROUP_FILL; cell.font = _GROUP_FONT
            elif is_note:
                cell.fill = _NOTE_FILL; cell.font = Font(name="Calibri", size=10, italic=True)
            elif has_con:
                cell.fill = _STOP_FILL; cell.font = Font(name="Calibri", size=10)
            elif is_opt:
                cell.fill = _OPT_FILL; cell.font = Font(name="Calibri", size=10, italic=True)
            else:
                cell.font = Font(name="Calibri", size=10)
        ws_s.row_dimensions[ri].height = 16

    col_w = {"type":30,"name":28,"label":65,"hint":40,"required":11,
             "relevant":60,"constraint":55,"constraint_message":50,
             "appearance":18,"default":18,"calculation":40}
    for ci, col in enumerate(used_s, 1):
        ws_s.column_dimensions[get_column_letter(ci)].width = col_w.get(col,20)
    ws_s.freeze_panes = "A2"
    if survey_rows:
        ws_s.auto_filter.ref = f"A1:{get_column_letter(len(used_s))}1"

    _write_header(ws_c, _CHOICES_COLS, _H_CHOICES)
    prev_list, toggle = None, True
    fa = PatternFill("solid", fgColor="E8F5E9")
    fb = PatternFill("solid", fgColor="FFFFFF")
    for ri, row in enumerate(choices_rows, 2):
        if row.get("list_name") != prev_list:
            toggle = not toggle; prev_list = row["list_name"]
        fill = fa if toggle else fb
        for ci, col in enumerate(_CHOICES_COLS, 1):
            val  = row.get(col,"")
            cell = ws_c.cell(row=ri, column=ci, value=str(val) if val else "")
            cell.font = Font(name="Calibri", size=10); cell.fill = fill
            cell.border = _BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=(col=="label"))
        ws_c.row_dimensions[ri].height = 15
    ws_c.column_dimensions["A"].width = 35
    ws_c.column_dimensions["B"].width = 22
    ws_c.column_dimensions["C"].width = 60
    ws_c.freeze_panes = "A2"
    if choices_rows:
        ws_c.auto_filter.ref = "A1:C1"

    used_t = [c for c in _SETTINGS_COLS if settings.get(c)]
    _write_header(ws_t, used_t, _H_SETTINGS)
    for ci, col in enumerate(used_t, 1):
        cell = ws_t.cell(row=2, column=ci, value=str(settings.get(col,"")))
        cell.font = Font(name="Calibri", size=11, bold=True)
        cell.alignment = Alignment(horizontal="left", vertical="center")
        cell.border = _BORDER
    ws_t.row_dimensions[2].height = 20
    for ci in range(1, len(used_t)+1):
        ws_t.column_dimensions[get_column_letter(ci)].width = 32

    out = io.BytesIO(); wb.save(out); out.seek(0)
    return out.read()


# ── Public API ─────────────────────────────────────────────────────────────────

def convert_docx_to_xlsform(docx_path: str) -> tuple[bytes, dict]:
    from .docx_parser import parse_questionnaire
    qs = parse_questionnaire(docx_path)
    return convert_json_to_xlsform({"_meta": {"source": Path(docx_path).name}, "questions": qs})


def convert_parsed_json_to_xlsform(json_path: str) -> tuple[bytes, dict]:
    data = json.loads(Path(json_path).read_text(encoding="utf-8"))
    return convert_json_to_xlsform(data)


if __name__ == "__main__":
    import sys
    from dotenv import load_dotenv; load_dotenv()
    logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s")
    src = sys.argv[1] if len(sys.argv) > 1 else "/tmp/files2/questionnaire_parsed.json"
    out = sys.argv[2] if len(sys.argv) > 2 else "/tmp/test_output.xlsx"
    xlsx, notes = convert_parsed_json_to_xlsform(src) if src.endswith(".json") \
           else convert_docx_to_xlsform(src)
    Path(out).write_bytes(xlsx)
    print(f"✓ {out} ({len(xlsx):,} bytes)")
    if notes["fallback_questions"]:
        print(f"  Fallback: {len(notes['fallback_questions'])} pertanyaan")
    if notes["placeholder_choices"]:
        print(f"  Placeholder: {len(notes['placeholder_choices'])} list")
