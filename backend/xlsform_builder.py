"""
xlsform_builder.py
Converts XLSForm JSON data into a properly formatted .xlsx file
compatible with KoboToolbox / ODK.
"""

import io
import logging
from typing import Any

import openpyxl
from openpyxl.styles import (
    Font,
    PatternFill,
    Alignment,
    Border,
    Side,
)
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

# XLSForm column definitions (order matters)
SURVEY_COLUMNS = [
    "type",
    "name",
    "label",
    "hint",
    "required",
    "relevant",
    "constraint",
    "constraint_message",
    "appearance",
    "calculation",
    "default",
    "read_only",
    "repeat_count",
    "media::image",
    "media::audio",
    "body::accuracyThreshold",
    "parameters",
]

CHOICES_COLUMNS = [
    "list_name",
    "name",
    "label",
    "image",
]

SETTINGS_COLUMNS = [
    "form_title",
    "form_id",
    "version",
    "default_language",
    "instance_name",
    "submission_url",
    "public_key",
    "auto_send",
    "auto_delete",
    "allow_choice_duplicates",
]

# Header style
HEADER_FILL = PatternFill("solid", fgColor="1F4E79")  # dark blue
HEADER_FONT = Font(color="FFFFFF", bold=True, name="Calibri", size=11)

# Group row style
GROUP_FILL = PatternFill("solid", fgColor="D6E4F0")
GROUP_FONT = Font(bold=True, name="Calibri", size=10, italic=True)

# Note/calculate row style
NOTE_FILL = PatternFill("solid", fgColor="F2F2F2")

# Required column style
REQUIRED_FILL = PatternFill("solid", fgColor="FFF2CC")

# Thin border
THIN_SIDE = Side(style="thin", color="BFBFBF")
THIN_BORDER = Border(
    left=THIN_SIDE,
    right=THIN_SIDE,
    top=THIN_SIDE,
    bottom=THIN_SIDE,
)

# Choice header style
CHOICE_HEADER_FILL = PatternFill("solid", fgColor="145A32")
CHOICE_HEADER_FONT = Font(color="FFFFFF", bold=True, name="Calibri", size=11)

# Settings header style
SETTINGS_HEADER_FILL = PatternFill("solid", fgColor="4A235A")
SETTINGS_HEADER_FONT = Font(color="FFFFFF", bold=True, name="Calibri", size=11)


def build_xlsform_from_json(data: dict) -> bytes:
    """
    Build an XLSForm .xlsx file from parsed JSON data.
    Returns raw bytes of the Excel file.
    """
    survey_rows: list = data.get("survey", [])
    choices_rows: list = data.get("choices", [])
    settings_dict: dict = data.get("settings", {})

    wb = openpyxl.Workbook()

    # Create sheets
    ws_survey = wb.active
    ws_survey.title = "survey"
    ws_choices = wb.create_sheet("choices")
    ws_settings = wb.create_sheet("settings")

    _write_survey_sheet(ws_survey, survey_rows)
    _write_choices_sheet(ws_choices, choices_rows)
    _write_settings_sheet(ws_settings, settings_dict)

    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    logger.info("XLSForm Excel file built successfully.")
    return output.read()


def _write_survey_sheet(ws, rows: list[dict]):
    """Write the survey sheet."""
    # Determine which columns are actually used (keep defined order)
    used_cols = [c for c in SURVEY_COLUMNS if any(row.get(c) for row in rows)]
    # Always include core columns
    core = ["type", "name", "label"]
    for c in core:
        if c not in used_cols:
            used_cols.insert(core.index(c), c)

    # Write header row
    for col_idx, col_name in enumerate(used_cols, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER

    ws.row_dimensions[1].height = 24

    # Write data rows
    for row_idx, row_data in enumerate(rows, 2):
        row_type = str(row_data.get("type", "")).lower().strip()
        is_group = row_type in ("begin_group", "end_group", "begin_repeat", "end_repeat")
        is_note_calc = row_type in ("note", "calculate", "start", "end")

        for col_idx, col_name in enumerate(used_cols, 1):
            value = row_data.get(col_name, "")
            if value is None:
                value = ""
            cell = ws.cell(row=row_idx, column=col_idx, value=str(value) if value else "")
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = THIN_BORDER

            # Styling by row type
            if is_group:
                cell.fill = GROUP_FILL
                cell.font = GROUP_FONT
            elif is_note_calc:
                cell.fill = NOTE_FILL
                cell.font = Font(name="Calibri", size=10, italic=True)
            else:
                cell.font = Font(name="Calibri", size=10)

            # Highlight required column if "yes"
            if col_name == "required" and str(value).lower() == "yes":
                cell.fill = REQUIRED_FILL
                cell.font = Font(name="Calibri", size=10, bold=True, color="7B3F00")

        ws.row_dimensions[row_idx].height = 18

    # Set column widths
    col_widths = {
        "type": 28,
        "name": 30,
        "label": 55,
        "hint": 45,
        "required": 12,
        "relevant": 55,
        "constraint": 55,
        "constraint_message": 55,
        "appearance": 18,
        "calculation": 55,
        "default": 18,
    }
    for col_idx, col_name in enumerate(used_cols, 1):
        width = col_widths.get(col_name, 20)
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Freeze top row
    ws.freeze_panes = "A2"

    # Auto-filter
    if rows:
        last_col = get_column_letter(len(used_cols))
        ws.auto_filter.ref = f"A1:{last_col}1"


def _write_choices_sheet(ws, rows: list[dict]):
    """Write the choices sheet."""
    used_cols = ["list_name", "name", "label"]
    if any(row.get("image") for row in rows):
        used_cols.append("image")

    # Header
    for col_idx, col_name in enumerate(used_cols, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = CHOICE_HEADER_FONT
        cell.fill = CHOICE_HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER

    ws.row_dimensions[1].height = 24

    # Alternate row colors for readability
    prev_list = None
    color_a = PatternFill("solid", fgColor="EAF4FB")
    color_b = PatternFill("solid", fgColor="FFFFFF")
    use_color_a = True

    for row_idx, row_data in enumerate(rows, 2):
        current_list = row_data.get("list_name", "")
        if current_list != prev_list:
            use_color_a = not use_color_a
            prev_list = current_list

        fill = color_a if use_color_a else color_b

        for col_idx, col_name in enumerate(used_cols, 1):
            value = row_data.get(col_name, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=str(value) if value else "")
            cell.font = Font(name="Calibri", size=10)
            cell.fill = fill
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=col_name == "label")

        ws.row_dimensions[row_idx].height = 16

    # Column widths
    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 55
    if "image" in used_cols:
        ws.column_dimensions["D"].width = 30

    ws.freeze_panes = "A2"

    if rows:
        last_col = get_column_letter(len(used_cols))
        ws.auto_filter.ref = f"A1:{last_col}1"


def _write_settings_sheet(ws, settings: dict):
    """Write the settings sheet."""
    # Use only known columns that have values
    used_cols = [c for c in SETTINGS_COLUMNS if settings.get(c)]
    if not used_cols:
        used_cols = ["form_title", "form_id", "version", "default_language"]

    # Header row
    for col_idx, col_name in enumerate(used_cols, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = SETTINGS_HEADER_FONT
        cell.fill = SETTINGS_HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER

    # Values row
    for col_idx, col_name in enumerate(used_cols, 1):
        value = settings.get(col_name, "")
        cell = ws.cell(row=2, column=col_idx, value=str(value) if value else "")
        cell.font = Font(name="Calibri", size=11, bold=True)
        cell.alignment = Alignment(horizontal="left", vertical="center")
        cell.border = THIN_BORDER

    ws.row_dimensions[1].height = 24
    ws.row_dimensions[2].height = 20

    for col_idx in range(1, len(used_cols) + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 30
